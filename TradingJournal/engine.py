import json
import os
from openpyxl import load_workbook

EXCEL_FILE = "uploaded.xlsx"
DB_FILE = "trades.json"


def load_db():
    if not os.path.exists(DB_FILE):
        return []
    with open(DB_FILE, "r") as f:
        return json.load(f)


def save_db(data):
    with open(DB_FILE, "w") as f:
        json.dump(data, f, indent=4)


def parse_excel():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    rows = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue

        if not header_found:
            if row[0] == "Time" and row[1] == "Position":
                header_found = True
            continue

        try:
            rows.append({
                "position_id": int(row[1]),
                "symbol": str(row[2]),
                "type": str(row[3]).upper(),
                "volume": float(row[4]) if row[4] else 0.0,
                "entry_price": float(row[5]) if row[5] else 0.0,
                "sl": float(row[6]) if row[6] else 0.0,
                "tp": float(row[7]) if row[7] else 0.0,
                "exit_price": float(row[9]) if row[9] else 0.0,
                "profit": float(row[12]) if row[12] else 0.0,
                "time": str(row[0]),
            })
        except:
            continue

    return rows


def build_trades(rows):
    grouped = {}

    for r in rows:
        pid = r["position_id"]

        if pid not in grouped:
            grouped[pid] = {
                "position_id": pid,
                "symbol": r["symbol"],
                "type": r["type"],
                "volume": 0.0,
                "entry_price": r["entry_price"],
                "exit_price": r["exit_price"],
                "sl": r["sl"],
                "tp": r["tp"],
                "profit": 0.0,
                "open_time": r["time"],
                "close_time": r["time"],
            }

        grouped[pid]["volume"] += r["volume"]
        grouped[pid]["profit"] += r["profit"]

    final = []

    for t in grouped.values():
        if abs(t["profit"]) < 0.0001:
            continue

        t["journal"] = {
            "pre_trade": "",
            "post_trade": "",
            "risk_reward": "",
            "emotions": "",
            "lessons": "",
            "tags": "",
            "rating": 5
        }

        final.append(t)

    return final


def sync_trades():
    db = load_db()
    db_map = {t["position_id"]: t for t in db}

    new_trades = build_trades(parse_excel())

    for t in new_trades:
        pid = t["position_id"]

        if pid in db_map:
            existing = db_map[pid]
            t["journal"] = existing.get("journal", t["journal"])
            db_map[pid] = t
        else:
            db_map[pid] = t

    db = list(db_map.values())
    save_db(db)

    return db